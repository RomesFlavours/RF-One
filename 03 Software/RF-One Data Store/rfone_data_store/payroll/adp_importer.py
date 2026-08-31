"""ADP `Payroll Detail` Excel importer (TASK_PAYROLL_001 §21-30).

ADP is only the current provider (`Payroll Provider Result.md`, "Provider
boundary") — nothing here is imported into canonical Payroll semantics as an
ADP-specific field; every table this module writes to is provider-generic
(`PayrollEarningFact.earning_type` etc. are free strings normalized from
whatever label the source actually used).

This module only reads a local `.xlsx` file with `openpyxl` and writes to
the RF-One database via SQLAlchemy — no network/API access of any kind is
used or required here. `Payroll Detail` reports are treated as source
documents: never modified, never re-saved. This is the manual/local-file
acquisition path — fully preserved as a valid acquisition adapter and
production fallback (TASK_PAYROLL_003, `rfone_data_store/payroll/
acquisition.py`), not the only way payroll results may reach RF-One.

Entry points:

- `dry_run_import(...)` / `dry_run_parsed_import(...)` — read-only. Parses
  the workbook (from a file, or from an already-parsed result any
  acquisition adapter produced), resolves Employee identity against
  already-confirmed mappings plus exact-unique name-key matching, and
  returns an aggregate-only `DryRunSummary`. Writes nothing.
- `persist_import(...)` — the file-based entry point: parses a local file,
  then delegates to `persist_parsed_import(...)`.
- `persist_parsed_import(...)` — the acquisition-method-independent core:
  writes `PayrollRun`/`EmployeePayrollResult`/fact rows (and
  `PayrollProviderEmployeeIdentity` mapping rows) onto the given Session for
  an already-parsed result, regardless of how it was acquired. The caller
  commits or rolls back, matching this repository's existing
  `calculate_tips.py` convention. Every acquisition adapter in
  `rfone_data_store/payroll/acquisition.py` normalizes into a
  `ParsedPayrollDetail` and calls this same function — no acquisition path
  has its own persistence/idempotency logic.

Idempotency (task §29): re-importing the exact same content (same SHA-256,
regardless of transport) for the same (source_system_id, restaurant_id) is
detected via `payroll_import_runs`' unique constraint and is a safe no-op —
it reuses the prior `PayrollImportRun`/`PayrollRun` rather than duplicating
anything. A *different* result for the same scope is never assumed to be a
correction of a specific prior run — that requires the caller to pass
`supersedes_import_run_id` explicitly (task §29, "Design a clear
correction/replacement/reconciliation behavior").
"""

from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from pathlib import Path

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from .. import models as m
from . import payment_execution as pe

UTC = timezone.utc

EMPLOYEE_NAME_HEADER = "Employee Name"
EARNING_GROUP_RE = re.compile(r"^Earning\s+(\d+)$", re.IGNORECASE)
PAYMENT_MARKER_RE = re.compile(r"^Payment\s+(\d+)$", re.IGNORECASE)
PAYMENT_CHECK_DATE_RE = re.compile(r"^Payment\s+(\d+)\s+Check Date$", re.IGNORECASE)
PAYMENT_REFERENCE_RE = re.compile(r"^Payment\s+(\d+)\s+Transaction ID or Check #$", re.IGNORECASE)
PAYMENT_AMOUNT_RE = re.compile(r"^Payment\s+(\d+)\s+Amount$", re.IGNORECASE)
EMPLOYER_LIABILITY_RE = re.compile(r".*-ER$", re.IGNORECASE)
CHECK_DATES_FROM_RE = re.compile(
    r"Check Dates From:\s*([\d/]+)\s*-\s*(.+)$", re.IGNORECASE
)
TOTAL_ROW_MARKER = "total"  # case-insensitive substring: real Employee Names never contain it


# ---------------------------------------------------------------------------
# Parsing (pure — no DB access)
# ---------------------------------------------------------------------------


@dataclass
class ParsedEarningLine:
    source_label: str
    earning_type: str
    quantity: Decimal | None
    unit: str | None
    rate_minor: int | None
    amount_minor: int
    paid_to_employee: bool
    excluded_from_taxable_wages: bool | None
    sequence: int


@dataclass
class ParsedLiabilityLine:
    source_label: str
    liability_type: str
    amount_minor: int


@dataclass
class ParsedPaymentLine:
    payment_method: str | None
    check_date: datetime | None
    reference: str | None
    amount_minor: int
    sequence: int


@dataclass
class ParsedEmployeeRow:
    row_number: int
    name_raw: str
    ssn_masked: str | None
    pay_frequency: str | None
    department: str | None
    earnings: list[ParsedEarningLine] = field(default_factory=list)
    liabilities: list[ParsedLiabilityLine] = field(default_factory=list)
    payments: list[ParsedPaymentLine] = field(default_factory=list)
    reported_total_employer_liability_minor: int | None = None


@dataclass
class ParsedPayrollDetail:
    header_pay_date: datetime | None
    header_provider_reference: str | None
    employees: list[ParsedEmployeeRow]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def to_minor_units(value) -> int | None:
    """Convert a source dollar amount (float/int/str, possibly blank) to
    integer minor units (cents) — never floating point arithmetic on the
    stored value itself (matches this schema's existing money convention)."""
    if value is None or value == "":
        return None
    d = Decimal(str(value))
    return int((d * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _clean(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_earning_label(raw_label: str) -> tuple[str, bool, bool | None]:
    """Parse the provider's own "* Items Not Paid To Employee" / "** Items
    Not Paid To Employee and Excluded From Some Wages" convention
    (`Payroll Provider Result.md`) off the end of a raw earning label, and
    normalize the remaining text into a stable `earning_type`.

    Returns (earning_type, paid_to_employee, excluded_from_taxable_wages).
    An earning label this importer has never seen before is normalized the
    same generic way — never rejected, never requiring a code change
    (task §26).
    """
    text = raw_label.strip()
    excluded_from_taxable_wages: bool | None = None
    paid_to_employee = True

    if text.endswith("**"):
        paid_to_employee = False
        excluded_from_taxable_wages = True
        text = text[:-2].strip()
    elif text.endswith("*"):
        paid_to_employee = False
        excluded_from_taxable_wages = False
        text = text[:-1].strip()

    earning_type = re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_").upper()
    return earning_type or "UNKNOWN", paid_to_employee, excluded_from_taxable_wages


def normalize_liability_label(raw_label: str) -> str:
    text = raw_label.strip()
    return re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_").upper() or "UNKNOWN"


def _parse_header(ws) -> tuple[datetime | None, str | None]:
    pay_date: datetime | None = None
    reference: str | None = None
    for row in ws.iter_rows(min_row=1, max_row=6, max_col=4):
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            match = CHECK_DATES_FROM_RE.search(cell.value)
            if match:
                date_text, ref_text = match.group(1), match.group(2).strip()
                try:
                    pay_date = datetime.strptime(date_text, "%m/%d/%Y").replace(tzinfo=UTC)
                except ValueError:
                    pay_date = None
                reference = ref_text or None
    return pay_date, reference


def _find_header_row(ws) -> int:
    for row in ws.iter_rows(min_row=1, max_row=15):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == EMPLOYEE_NAME_HEADER:
                return cell.row
    raise ValueError(f"Could not locate the '{EMPLOYEE_NAME_HEADER}' header row in the workbook")


@dataclass
class _ColumnMap:
    name_col: int
    ssn_col: int | None
    pay_frequency_col: int | None
    department_col: int | None
    earning_groups: list[tuple[int, int, int, int]]  # (label_col, hours_col, rate_col, amount_col)
    liability_cols: list[tuple[int, str]]  # (col, raw_label)
    payment_groups: dict[int, dict[str, int]]  # seq -> {method,date,ref,amount: col}
    total_employer_liability_col: int | None


def _build_column_map(ws, header_row: int) -> _ColumnMap:
    headers: dict[int, str] = {}
    for cell in ws[header_row]:
        if cell.value is not None:
            headers[cell.column] = str(cell.value).strip()

    def find_single(label: str) -> int | None:
        for col, text in headers.items():
            if text == label:
                return col
        return None

    name_col = find_single(EMPLOYEE_NAME_HEADER)
    if name_col is None:
        raise ValueError(f"'{EMPLOYEE_NAME_HEADER}' column not found on header row {header_row}")

    earning_groups: list[tuple[int, int, int, int]] = []
    payment_groups: dict[int, dict[str, int]] = {}
    liability_cols: list[tuple[int, str]] = []

    for col in sorted(headers):
        text = headers[col]
        earning_match = EARNING_GROUP_RE.match(text)
        if earning_match:
            hours_col, rate_col, amount_col = col + 1, col + 2, col + 3
            for offset_col, expected in ((hours_col, "hours"), (rate_col, "rate"), (amount_col, "amount")):
                offset_text = headers.get(offset_col, "")
                if expected not in offset_text.lower():
                    raise ValueError(
                        f"Expected an '{expected}' column at position {offset_col} following "
                        f"'{text}' (col {col}), found {offset_text!r} instead — workbook layout "
                        "does not match the expected 'Earning N / Hours / Rate / Amount' structure."
                    )
            earning_groups.append((col, hours_col, rate_col, amount_col))
            continue

        payment_marker = PAYMENT_MARKER_RE.match(text)
        if payment_marker:
            seq = int(payment_marker.group(1))
            payment_groups.setdefault(seq, {})["method"] = col
            continue
        check_date_match = PAYMENT_CHECK_DATE_RE.match(text)
        if check_date_match:
            seq = int(check_date_match.group(1))
            payment_groups.setdefault(seq, {})["date"] = col
            continue
        reference_match = PAYMENT_REFERENCE_RE.match(text)
        if reference_match:
            seq = int(reference_match.group(1))
            payment_groups.setdefault(seq, {})["ref"] = col
            continue
        amount_match = PAYMENT_AMOUNT_RE.match(text)
        if amount_match:
            seq = int(amount_match.group(1))
            payment_groups.setdefault(seq, {})["amount"] = col
            continue

        if EMPLOYER_LIABILITY_RE.match(text):
            liability_cols.append((col, text))

    return _ColumnMap(
        name_col=name_col,
        ssn_col=find_single("SSN"),
        pay_frequency_col=find_single("Pay Frequency"),
        department_col=find_single("Department"),
        earning_groups=earning_groups,
        liability_cols=liability_cols,
        payment_groups=payment_groups,
        total_employer_liability_col=find_single("Total Employer Liability"),
    )


def parse_payroll_detail_workbook(path: Path) -> ParsedPayrollDetail:
    """Parse an ADP `Payroll Detail` export from a local file. Pure/read-only:
    the workbook itself is never modified (task §36)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        return _parse_loaded_workbook(wb)
    finally:
        wb.close()


def parse_payroll_detail_workbook_bytes(data: bytes) -> ParsedPayrollDetail:
    """Parse an ADP `Payroll Detail` export already held in memory
    (TASK_PAYROLL_003) — the same parser `parse_payroll_detail_workbook`
    uses, for acquisition paths that never write a local file (e.g. an SFTP
    download). Identical parsing logic either way; only the source of the
    bytes differs."""
    wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
    try:
        return _parse_loaded_workbook(wb)
    finally:
        wb.close()


def _parse_loaded_workbook(wb) -> ParsedPayrollDetail:
    ws = wb[wb.sheetnames[0]]

    header_pay_date, header_reference = _parse_header(ws)
    header_row = _find_header_row(ws)
    columns = _build_column_map(ws, header_row)

    employees: list[ParsedEmployeeRow] = []
    for offset, row in enumerate(ws.iter_rows(min_row=header_row + 1)):
        row_number = header_row + 1 + offset
        if not row:
            continue
        name_cell = row[columns.name_col - 1].value if columns.name_col - 1 < len(row) else None
        name_raw = _clean(name_cell)
        if not name_raw:
            continue
        if TOTAL_ROW_MARKER in name_raw.lower():
            # Summary rows ("Company Total and Employee Count", "Pay
            # Frequency Total and Employee Count") — never Employee facts.
            break

        def cell_value(col: int):
            idx = col - 1
            return row[idx].value if idx < len(row) else None

        earnings: list[ParsedEarningLine] = []
        for sequence, (label_col, hours_col, rate_col, amount_col) in enumerate(
            columns.earning_groups, start=1
        ):
            label = _clean(cell_value(label_col))
            amount = to_minor_units(cell_value(amount_col))
            if not label or amount is None:
                continue
            earning_type, paid_to_employee, excluded = normalize_earning_label(label)
            hours_val = cell_value(hours_col)
            quantity = Decimal(str(hours_val)) if hours_val not in (None, "") else None
            rate_minor = to_minor_units(cell_value(rate_col))
            earnings.append(
                ParsedEarningLine(
                    source_label=label,
                    earning_type=earning_type,
                    quantity=quantity,
                    unit="HOURS" if quantity is not None else None,
                    rate_minor=rate_minor,
                    amount_minor=amount,
                    paid_to_employee=paid_to_employee,
                    excluded_from_taxable_wages=excluded,
                    sequence=sequence,
                )
            )

        liabilities: list[ParsedLiabilityLine] = []
        for col, raw_label in columns.liability_cols:
            amount = to_minor_units(cell_value(col))
            if amount is None:
                continue
            liabilities.append(
                ParsedLiabilityLine(
                    source_label=raw_label,
                    liability_type=normalize_liability_label(raw_label),
                    amount_minor=amount,
                )
            )

        payments: list[ParsedPaymentLine] = []
        for seq in sorted(columns.payment_groups):
            group = columns.payment_groups[seq]
            amount_col = group.get("amount")
            if amount_col is None:
                continue
            amount = to_minor_units(cell_value(amount_col))
            if amount is None:
                continue
            method = _clean(cell_value(group["method"])) if "method" in group else None
            date_val = cell_value(group["date"]) if "date" in group else None
            check_date: datetime | None = None
            if isinstance(date_val, datetime):
                check_date = date_val if date_val.tzinfo else date_val.replace(tzinfo=UTC)
            elif isinstance(date_val, str) and date_val.strip():
                try:
                    check_date = datetime.strptime(date_val.strip(), "%m/%d/%Y").replace(tzinfo=UTC)
                except ValueError:
                    check_date = None
            reference = _clean(cell_value(group["ref"])) if "ref" in group else None
            payments.append(
                ParsedPaymentLine(
                    payment_method=method,
                    check_date=check_date,
                    reference=reference,
                    amount_minor=amount,
                    sequence=seq,
                )
            )

        reported_total_liability = (
            to_minor_units(cell_value(columns.total_employer_liability_col))
            if columns.total_employer_liability_col
            else None
        )

        employees.append(
            ParsedEmployeeRow(
                row_number=row_number,
                name_raw=name_raw,
                ssn_masked=_clean(cell_value(columns.ssn_col)) if columns.ssn_col else None,
                pay_frequency=_clean(cell_value(columns.pay_frequency_col))
                if columns.pay_frequency_col
                else None,
                department=_clean(cell_value(columns.department_col))
                if columns.department_col
                else None,
                earnings=earnings,
                liabilities=liabilities,
                payments=payments,
                reported_total_employer_liability_minor=reported_total_liability,
            )
        )

    return ParsedPayrollDetail(
        header_pay_date=header_pay_date,
        header_provider_reference=header_reference,
        employees=employees,
    )


# ---------------------------------------------------------------------------
# Employee identity mapping — exact, structural, never fuzzy
# (Payroll Provider Result.md, "Employee mapping")
# ---------------------------------------------------------------------------

RESOLVED = "RESOLVED"
UNRESOLVED = "UNRESOLVED"
AMBIGUOUS = "AMBIGUOUS"


def adp_name_key(raw_name: str) -> tuple[str, str]:
    """`"Last, First Middle"` -> (first, last), lowercased. Only the first
    token after the comma is used as the first name — a middle name/initial
    on the ADP side is dropped deterministically, never approximated."""
    last_part, _, rest = raw_name.partition(",")
    last = last_part.strip().lower()
    rest_tokens = rest.strip().split()
    first = rest_tokens[0].lower() if rest_tokens else ""
    return first, last


def employee_display_name_key(display_name: str) -> tuple[str, str]:
    """`"First Last"` (or `"First Middle Last"`) -> (first, last-and-rest),
    lowercased. The first token is the first name; everything else is
    joined as the last name, so a compound last name on the RF-One side
    matches an ADP `"Last-Part, First"` row without a fuzzy comparison."""
    tokens = display_name.strip().split()
    if not tokens:
        return "", ""
    first = tokens[0].lower()
    last = " ".join(tokens[1:]).lower()
    return first, last


@dataclass
class MappingResolution:
    external_employee_key: str
    status: str
    employee_id: int | None
    resolution_method: str | None


def _restaurant_location_ids(session: Session, restaurant_id: int) -> set[int]:
    rows = session.scalars(
        select(m.RestaurantLocation.location_id).where(
            m.RestaurantLocation.restaurant_id == restaurant_id
        )
    ).all()
    return set(rows)


def resolve_employee_mappings(
    session: Session, *, source_system_id: int, restaurant_id: int, rows: list[ParsedEmployeeRow]
) -> dict[str, MappingResolution]:
    """Resolve each distinct ADP name key to an RF-One Employee.

    Order of resolution, per row:
    1. An existing `PayrollProviderEmployeeIdentity` row for this
       (source_system_id, restaurant_id, external_employee_key) — reused
       as-is, whatever its status (a human may have already confirmed or
       rejected it).
    2. Otherwise, exact-unique structural name-key matching against current
       Employees in this Restaurant's scope. Exactly one match -> RESOLVED.
       Zero or more-than-one match -> UNRESOLVED/AMBIGUOUS. Never a
       similarity/fuzzy match (task §28).

    Returns a dict keyed by `external_employee_key` — never persists
    anything; `persist_import` is responsible for writing new mapping rows.
    """
    location_ids = _restaurant_location_ids(session, restaurant_id)
    employees = session.execute(
        select(m.Employee.id, m.Employee.display_name).where(
            m.Employee.location_id.in_(location_ids) if location_ids else False,
            m.Employee.display_name.is_not(None),
        )
    ).all()

    by_key: dict[tuple[str, str], list[int]] = {}
    for employee_id, display_name in employees:
        key = employee_display_name_key(display_name)
        by_key.setdefault(key, []).append(employee_id)

    existing_mappings = {
        row.external_employee_key: row
        for row in session.scalars(
            select(m.PayrollProviderEmployeeIdentity).where(
                m.PayrollProviderEmployeeIdentity.source_system_id == source_system_id,
                m.PayrollProviderEmployeeIdentity.restaurant_id == restaurant_id,
            )
        ).all()
    }

    resolutions: dict[str, MappingResolution] = {}
    seen_keys: set[str] = set()
    for parsed_row in rows:
        first, last = adp_name_key(parsed_row.name_raw)
        external_key = f"{first}:{last}"
        if external_key in seen_keys:
            continue
        seen_keys.add(external_key)

        existing = existing_mappings.get(external_key)
        if existing is not None:
            resolutions[external_key] = MappingResolution(
                external_employee_key=external_key,
                status=existing.mapping_status,
                employee_id=existing.employee_id,
                resolution_method=existing.resolution_method,
            )
            continue

        candidates = by_key.get((first, last), [])
        if len(candidates) == 1:
            resolutions[external_key] = MappingResolution(
                external_employee_key=external_key,
                status=RESOLVED,
                employee_id=candidates[0],
                resolution_method="EXACT_NAME_KEY_UNIQUE_MATCH",
            )
        elif len(candidates) > 1:
            resolutions[external_key] = MappingResolution(
                external_employee_key=external_key,
                status=AMBIGUOUS,
                employee_id=None,
                resolution_method=None,
            )
        else:
            resolutions[external_key] = MappingResolution(
                external_employee_key=external_key,
                status=UNRESOLVED,
                employee_id=None,
                resolution_method=None,
            )

    return resolutions


# ---------------------------------------------------------------------------
# Dry-run summary (task §35) — aggregate-only, never an Employee name
# ---------------------------------------------------------------------------


@dataclass
class DryRunSummary:
    employees_represented: int
    pay_dates: list[str]
    earning_line_count: int
    reportable_tip_line_count: int
    employer_liability_line_count: int
    payment_fact_count: int
    total_employer_paid_earnings_minor: int
    total_employer_liabilities_minor: int
    total_payroll_employer_cost_minor: int
    total_employee_payment_amount_minor: int
    unresolved_employee_mapping_count: int
    ambiguous_employee_mapping_count: int
    unparsed_source_labels: list[str]


def build_dry_run_summary(
    parsed: ParsedPayrollDetail, mappings: dict[str, MappingResolution]
) -> DryRunSummary:
    pay_dates: set[str] = set()
    earning_count = 0
    tip_count = 0
    liability_count = 0
    payment_count = 0
    total_paid_earnings = 0
    total_liabilities = 0
    total_payment = 0
    unparsed_labels: set[str] = set()

    for row in parsed.employees:
        for payment in row.payments:
            if payment.check_date is not None:
                pay_dates.add(payment.check_date.date().isoformat())
            payment_count += 1
            total_payment += payment.amount_minor
        for earning in row.earnings:
            earning_count += 1
            if earning.earning_type == "UNKNOWN":
                unparsed_labels.add(earning.source_label)
            if "TIP" in earning.earning_type:
                tip_count += 1
            if earning.paid_to_employee:
                total_paid_earnings += earning.amount_minor
        for liability in row.liabilities:
            liability_count += 1
            total_liabilities += liability.amount_minor

    unresolved = sum(1 for r in mappings.values() if r.status == UNRESOLVED)
    ambiguous = sum(1 for r in mappings.values() if r.status == AMBIGUOUS)

    if parsed.header_pay_date is not None:
        pay_dates.add(parsed.header_pay_date.date().isoformat())

    return DryRunSummary(
        employees_represented=len(parsed.employees),
        pay_dates=sorted(pay_dates),
        earning_line_count=earning_count,
        reportable_tip_line_count=tip_count,
        employer_liability_line_count=liability_count,
        payment_fact_count=payment_count,
        total_employer_paid_earnings_minor=total_paid_earnings,
        total_employer_liabilities_minor=total_liabilities,
        total_payroll_employer_cost_minor=total_paid_earnings + total_liabilities,
        total_employee_payment_amount_minor=total_payment,
        unresolved_employee_mapping_count=unresolved,
        ambiguous_employee_mapping_count=ambiguous,
        unparsed_source_labels=sorted(unparsed_labels),
    )


def dry_run_import(
    session: Session, *, source_system_id: int, restaurant_id: int, file_path: Path
) -> DryRunSummary:
    """Read-only: parses the workbook and resolves Employee mapping against
    already-confirmed mappings plus exact-unique matching. Writes nothing."""
    parsed = parse_payroll_detail_workbook(file_path)
    return dry_run_parsed_import(session, source_system_id=source_system_id, restaurant_id=restaurant_id, parsed=parsed)


def dry_run_parsed_import(
    session: Session, *, source_system_id: int, restaurant_id: int, parsed: ParsedPayrollDetail
) -> DryRunSummary:
    """Read-only, acquisition-method-independent (TASK_PAYROLL_003): the
    same summary `dry_run_import` produces, for a result already parsed
    from any acquisition adapter. Writes nothing."""
    mappings = resolve_employee_mappings(
        session, source_system_id=source_system_id, restaurant_id=restaurant_id, rows=parsed.employees
    )
    return build_dry_run_summary(parsed, mappings)


# ---------------------------------------------------------------------------
# Persistence — idempotent, provenance-aware (task §29)
# ---------------------------------------------------------------------------


@dataclass
class PersistResult:
    import_run_id: int
    payroll_run_id: int | None
    created: bool
    employees_persisted: int
    unresolved_employee_count: int
    ambiguous_employee_count: int
    issue_count: int


def persist_import(
    session: Session,
    *,
    source_system_id: int,
    restaurant_id: int,
    file_path: Path,
    period_start: datetime | None,
    period_end: datetime | None,
    run_type: str,
    payroll_schedule_id: int | None = None,
    pay_date_override: datetime | None = None,
    supersedes_import_run_id: int | None = None,
    payment_execution_provider: str | None = None,
) -> PersistResult:
    """Parse a local ADP `Payroll Detail` file and write it via
    `persist_parsed_import` (`acquisition_method="ADP_XLSX_FILE"`) — the
    manual/local-file acquisition path, fully preserved as a valid
    acquisition adapter and production fallback (TASK_PAYROLL_003). See
    `persist_parsed_import` for the shared persistence/idempotency contract
    every acquisition path (this one, and
    `rfone_data_store/payroll/acquisition.py`'s SFTP/API adapters) reuses
    unchanged.
    """
    file_hash = sha256_file(file_path)
    parsed = parse_payroll_detail_workbook(file_path)
    return persist_parsed_import(
        session,
        source_system_id=source_system_id,
        restaurant_id=restaurant_id,
        parsed=parsed,
        source_file_name=file_path.name,
        source_file_hash=file_hash,
        acquisition_method="ADP_XLSX_FILE",
        period_start=period_start,
        period_end=period_end,
        run_type=run_type,
        payroll_schedule_id=payroll_schedule_id,
        pay_date_override=pay_date_override,
        supersedes_import_run_id=supersedes_import_run_id,
        payment_execution_provider=payment_execution_provider,
    )


def persist_parsed_import(
    session: Session,
    *,
    source_system_id: int,
    restaurant_id: int,
    parsed: ParsedPayrollDetail,
    source_file_name: str,
    source_file_hash: str,
    acquisition_method: str,
    period_start: datetime | None,
    period_end: datetime | None,
    run_type: str,
    payroll_schedule_id: int | None = None,
    pay_date_override: datetime | None = None,
    supersedes_import_run_id: int | None = None,
    payment_execution_provider: str | None = None,
) -> PersistResult:
    """Resolve and write an already-parsed ADP `Payroll Detail` result —
    the acquisition-method-independent core every acquisition adapter
    shares (TASK_PAYROLL_003, `Payroll Result Acquisition.md`). Idempotent
    by content hash (task §29): if a `PayrollImportRun` already exists for
    this exact (source_system_id, restaurant_id, `source_file_hash`), it is
    reused unchanged and no new PayrollRun/fact rows are created — this
    holds regardless of how the bytes arrived (a local file, an SFTP
    download, a future API response), because idempotency is keyed on
    content, never on filename, path, or transport.

    A *different* result for the same scope is a distinct import — it is
    never merged into a prior run's history automatically. Pass
    `supersedes_import_run_id` (an explicit operator decision) to mark that
    this import corrects a specific prior one; the prior PayrollRun's
    `status` becomes `SUPERSEDED` and its `superseded_by_payroll_run_id` is
    set, but it is never deleted or rewritten (task §29).

    `payment_execution_provider` (TASK_PAYROLL_003 correction of
    TASK_PAYROLL_002): defaults to `None` — it is **never** inferred merely
    because the acquisition source is ADP. The Run's provider is resolved
    as: (1) this explicit argument, if given; else (2) the Restaurant's
    approved `PayrollExecutionConfiguration` valid at the Run's `pay_date`,
    if one exists; else (3) left unassigned (`NULL`) — a Restaurant with no
    explicit selection and no approved configuration simply has no assigned
    executor yet, exactly like an unconfigured `TipPolicy`. This importer
    never reassigns an already-assigned Run's provider (double-payment
    prevention — see `rfone_data_store/payroll/payment_execution.py`).
    """
    existing = session.scalars(
        select(m.PayrollImportRun).where(
            m.PayrollImportRun.source_system_id == source_system_id,
            m.PayrollImportRun.restaurant_id == restaurant_id,
            m.PayrollImportRun.source_file_hash == source_file_hash,
        )
    ).first()
    if existing is not None:
        issue_count = len(existing.issues)
        return PersistResult(
            import_run_id=existing.id,
            payroll_run_id=existing.payroll_run_id,
            created=False,
            employees_persisted=existing.employees_represented_count or 0,
            unresolved_employee_count=existing.unresolved_employee_count or 0,
            ambiguous_employee_count=0,
            issue_count=issue_count,
        )

    mappings = resolve_employee_mappings(
        session, source_system_id=source_system_id, restaurant_id=restaurant_id, rows=parsed.employees
    )

    pay_date = pay_date_override or parsed.header_pay_date
    if pay_date is None:
        raise ValueError(
            "No pay_date could be determined from the source workbook and none was supplied — "
            "pass pay_date_override explicitly (Payroll Schedule and Period.md: Pay Date is "
            "never guessed)."
        )

    resolved_provider = payment_execution_provider
    if resolved_provider is None:
        resolved_provider = pe.approved_provider_at(session, restaurant_id=restaurant_id, at=pay_date)

    payroll_run = m.PayrollRun(
        restaurant_id=restaurant_id,
        source_system_id=source_system_id,
        payroll_schedule_id=payroll_schedule_id,
        period_start=period_start,
        period_end=period_end,
        pay_date=pay_date,
        run_type=run_type,
        provider_reference=parsed.header_provider_reference,
        status="COMPLETE",
    )
    if resolved_provider is not None:
        pe.assign_payment_execution_provider(payroll_run, resolved_provider)
    session.add(payroll_run)
    session.flush()

    import_run = m.PayrollImportRun(
        restaurant_id=restaurant_id,
        source_system_id=source_system_id,
        payroll_run_id=payroll_run.id,
        source_file_name=source_file_name,
        source_file_hash=source_file_hash,
        acquisition_method=acquisition_method,
        supersedes_import_run_id=supersedes_import_run_id,
        mode="PERSIST",
        status="COMPLETE",
    )
    session.add(import_run)
    session.flush()

    if supersedes_import_run_id is not None:
        prior_import = session.get(m.PayrollImportRun, supersedes_import_run_id)
        if prior_import is not None and prior_import.payroll_run_id is not None:
            prior_run = session.get(m.PayrollRun, prior_import.payroll_run_id)
            if prior_run is not None:
                prior_run.status = "SUPERSEDED"
                prior_run.superseded_by_payroll_run_id = payroll_run.id

    employees_persisted = 0
    unresolved_count = 0
    ambiguous_count = 0

    for row in parsed.employees:
        first, last = adp_name_key(row.name_raw)
        external_key = f"{first}:{last}"
        resolution = mappings[external_key]

        identity = session.scalars(
            select(m.PayrollProviderEmployeeIdentity).where(
                m.PayrollProviderEmployeeIdentity.source_system_id == source_system_id,
                m.PayrollProviderEmployeeIdentity.restaurant_id == restaurant_id,
                m.PayrollProviderEmployeeIdentity.external_employee_key == external_key,
            )
        ).first()
        if identity is None:
            identity = m.PayrollProviderEmployeeIdentity(
                source_system_id=source_system_id,
                restaurant_id=restaurant_id,
                external_employee_key=external_key,
                external_display_reference=row.ssn_masked,
                employee_id=resolution.employee_id,
                mapping_status=resolution.status,
                resolution_method=resolution.resolution_method,
                resolved_at=datetime.now(UTC) if resolution.status == RESOLVED else None,
            )
            session.add(identity)
            session.flush()

        if identity.mapping_status != RESOLVED or identity.employee_id is None:
            if identity.mapping_status == AMBIGUOUS:
                ambiguous_count += 1
                issue_type = "AMBIGUOUS_EMPLOYEE_MAPPING"
                severity = "BLOCKING"
            else:
                unresolved_count += 1
                issue_type = "UNRESOLVED_EMPLOYEE_MAPPING"
                severity = "BLOCKING"
            session.add(
                m.PayrollImportIssue(
                    import_run_id=import_run.id,
                    issue_type=issue_type,
                    severity=severity,
                    details=(
                        f"Row {row.row_number}: external_employee_key={external_key!r} "
                        f"status={identity.mapping_status} — payroll results for this row were "
                        "not persisted. Requires explicit manual mapping confirmation."
                    ),
                )
            )
            continue

        result = m.EmployeePayrollResult(
            payroll_run_id=payroll_run.id,
            employee_id=identity.employee_id,
            source_pay_frequency_label=row.pay_frequency,
            source_note=f"Department: {row.department}" if row.department else None,
        )
        session.add(result)
        session.flush()

        for earning in row.earnings:
            session.add(
                m.PayrollEarningFact(
                    employee_payroll_result_id=result.id,
                    earning_type=earning.earning_type,
                    source_label=earning.source_label,
                    quantity=earning.quantity,
                    unit=earning.unit,
                    rate_minor=earning.rate_minor,
                    amount_minor=earning.amount_minor,
                    paid_to_employee=earning.paid_to_employee,
                    excluded_from_taxable_wages=earning.excluded_from_taxable_wages,
                    sequence=earning.sequence,
                )
            )
            if earning.earning_type == "UNKNOWN":
                session.add(
                    m.PayrollImportIssue(
                        import_run_id=import_run.id,
                        issue_type="UNPARSED_SOURCE_ROW",
                        severity="WARNING",
                        details=(
                            f"Row {row.row_number}: earning label {earning.source_label!r} "
                            "normalized to an empty/UNKNOWN earning_type."
                        ),
                    )
                )

        for liability in row.liabilities:
            session.add(
                m.PayrollEmployerLiabilityFact(
                    employee_payroll_result_id=result.id,
                    liability_type=liability.liability_type,
                    source_label=liability.source_label,
                    amount_minor=liability.amount_minor,
                )
            )

        for payment in row.payments:
            session.add(
                m.PayrollPaymentFact(
                    employee_payroll_result_id=result.id,
                    pay_date=payment.check_date or pay_date,
                    payment_method=payment.payment_method,
                    payment_amount_minor=payment.amount_minor,
                    provider_payment_reference=payment.reference,
                    sequence=payment.sequence,
                )
            )

        employees_persisted += 1

    import_run.employees_represented_count = employees_persisted
    import_run.unresolved_employee_count = unresolved_count + ambiguous_count
    if unresolved_count or ambiguous_count:
        import_run.status = "PARTIAL"

    session.flush()

    return PersistResult(
        import_run_id=import_run.id,
        payroll_run_id=payroll_run.id,
        created=True,
        employees_persisted=employees_persisted,
        unresolved_employee_count=unresolved_count,
        ambiguous_employee_count=ambiguous_count,
        issue_count=len(import_run.issues),
    )
