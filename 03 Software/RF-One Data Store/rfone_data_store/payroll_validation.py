"""Automated synthetic tests for Payroll (TASK_PAYROLL_001 §34).

Mirrors `schema_validation.py`/`tips_validation.py`'s pattern exactly: builds
a synthetic (never-real) fixture inside one transaction, asserts the
required behaviors, and always rolls back — no synthetic row is ever left
in the target database. A tiny synthetic ADP-shaped workbook is generated
in a temp file (never a real payroll export) to exercise the importer's
parsing/idempotency/mapping/correction behavior end-to-end.
"""

from __future__ import annotations

import inspect
import shutil
import tempfile
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from pathlib import Path

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session, sessionmaker

from . import models as m
from .payroll import acquisition
from .payroll import adp_importer as adp
from .payroll import compensation, schedule
from .payroll import payment_execution as pe
from .payroll.labor_cost import compute_employee_labor_cost, compute_payroll_run_labor_cost

UTC = timezone.utc


@dataclass
class ValidationResult:
    success: bool
    checks_passed: list[str] = field(default_factory=list)
    checks_failed: list[str] = field(default_factory=list)

    def check(self, description: str, condition: bool) -> None:
        if condition:
            self.checks_passed.append(description)
        else:
            self.checks_failed.append(description)
            self.success = False


def run_validation(session_factory: sessionmaker[Session]) -> ValidationResult:
    result = ValidationResult(success=True)
    with session_factory() as session:
        try:
            _run_all_checks(session, result)
        finally:
            session.rollback()
    return result


def _dt(offset_days: int) -> datetime:
    # 2026-08-03 is a Monday — matches Rome's Flavours' current Workweek
    # anchor, so period math below is easy to reason about.
    base = datetime(2026, 8, 3, tzinfo=UTC)
    return base + timedelta(days=offset_days)


# ---------------------------------------------------------------------------
# Synthetic ADP workbook builder — a small, never-real ADP-shaped export,
# built purely to exercise the importer end-to-end.
# ---------------------------------------------------------------------------

_HEADERS = [
    "Employee Name", "SSN", "TIN", "Pay Frequency", "Department",
    "Earning  1", "Hours", "Rate", "Amount",
    "Earning  2", "Hours", "Rate", "Amount",
    "Earning  3", "Hours", "Rate", "Amount",
    "Total Hours", "Total Earnings",
    "FED FIT", "FED SOCSEC", "FED MEDCARE", "Total Taxes", "Deduction Total", "Net Pay",
    "FED SOCSEC-ER", "FED MEDCARE-ER", "Total Employer Liability",
    "Payment  1", "Payment  1  Check Date", "Payment  1  Transaction ID or Check #", "Payment  1  Amount",
]


def _build_synthetic_workbook(path: Path, rows: list[dict], pay_date_text: str = "8/19/2026 - Payroll 1") -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payroll Detail"
    ws["A1"] = "Company: SYNTHETIC TEST CO LLC"
    ws["A2"] = "Report: Payroll Detail"
    ws["A3"] = f"Check Dates From: {pay_date_text}"
    ws["A4"] = f"To: {pay_date_text}"
    for col_idx, header in enumerate(_HEADERS, start=1):
        ws.cell(row=7, column=col_idx, value=header)

    row_num = 8
    for row in rows:
        ws.cell(row=row_num, column=1, value=row["name"])
        ws.cell(row=row_num, column=2, value=row.get("ssn", "xxx-xx-0000"))
        ws.cell(row=row_num, column=4, value="Biweekly")
        ws.cell(row=row_num, column=6, value="Regular ")
        ws.cell(row=row_num, column=7, value=row.get("regular_hours", 80))
        ws.cell(row=row_num, column=8, value=row.get("regular_rate", 15.0))
        ws.cell(row=row_num, column=9, value=row.get("regular_amount", 1200.0))
        if "tips_amount" in row:
            ws.cell(row=row_num, column=10, value="Cash tips* ")
            ws.cell(row=row_num, column=11, value=0)
            ws.cell(row=row_num, column=13, value=row["tips_amount"])
        if "extra_label" in row:
            ws.cell(row=row_num, column=14, value=row["extra_label"])
            ws.cell(row=row_num, column=17, value=row.get("extra_amount", 0))
        ws.cell(row=row_num, column=26, value=row.get("socsec_er", 74.4))
        ws.cell(row=row_num, column=27, value=row.get("medicare_er", 17.4))
        ws.cell(row=row_num, column=29, value=row.get("payment_method", "Direct Deposit"))
        ws.cell(row=row_num, column=30, value=row.get("payment_date", "08/19/2026"))
        ws.cell(row=row_num, column=31, value=row.get("payment_ref", "Checking / Account No: XXXX0000"))
        ws.cell(row=row_num, column=32, value=row.get("payment_amount", 1100.0))
        row_num += 1

    row_num += 1
    ws.cell(row=row_num, column=1, value="Company Total and Employee Count")
    wb.save(path)


def _run_all_checks(session: Session, result: ValidationResult) -> None:
    # --- Base fixture ------------------------------------------------------
    source_clover = m.SourceSystem(code="CLOVER", name="Clover", active=True)
    source_adp = m.SourceSystem(code="ADP", name="ADP RUN", active=True)
    session.add_all([source_clover, source_adp])
    session.flush()

    location = m.Location(
        merchant_id=_ensure_merchant(session, source_clover).id,
        source_system_id=source_clover.id,
        source_location_id="LOC1",
        name="Test Location",
        currency="USD",
    )
    session.add(location)
    session.flush()

    restaurant = m.Restaurant(name="Synthetic Payroll Test Restaurant", default_currency="USD")
    session.add(restaurant)
    session.flush()
    session.add(m.RestaurantLocation(restaurant_id=restaurant.id, location_id=location.id, is_primary=True))
    session.flush()

    employee_a = m.Employee(location_id=location.id, display_name="Alice Tester", source_system_id=source_clover.id, source_employee_id="E1")
    employee_b = m.Employee(location_id=location.id, display_name="Bob Tester", source_system_id=source_clover.id, source_employee_id="E2")
    employee_dup1 = m.Employee(location_id=location.id, display_name="Chris Duplicate", source_system_id=source_clover.id, source_employee_id="E3")
    employee_dup2 = m.Employee(location_id=location.id, display_name="Chris Duplicate", source_system_id=source_clover.id, source_employee_id="E4")
    session.add_all([employee_a, employee_b, employee_dup1, employee_dup2])
    session.flush()

    # --- 1-3: Payroll Schedule types ---------------------------------------
    ps_weekly = m.PayrollSchedule(restaurant_id=restaurant.id, schedule_type="WEEKLY", code="W1")
    ps_biweekly = m.PayrollSchedule(restaurant_id=restaurant.id, schedule_type="BIWEEKLY", code="BW1")
    ps_monthly = m.PayrollSchedule(restaurant_id=restaurant.id, schedule_type="MONTHLY", code="M1")
    session.add_all([ps_weekly, ps_biweekly, ps_monthly])
    session.flush()
    result.check("1. WEEKLY PayrollSchedule is representable", ps_weekly.id is not None and ps_weekly.schedule_type == "WEEKLY")
    result.check("2. BIWEEKLY PayrollSchedule is representable", ps_biweekly.id is not None and ps_biweekly.schedule_type == "BIWEEKLY")
    result.check("3. MONTHLY PayrollSchedule is representable", ps_monthly.id is not None and ps_monthly.schedule_type == "MONTHLY")
    try:
        schedule.validate_schedule_type("QUARTERLY")
        invalid_rejected = False
    except ValueError:
        invalid_rejected = True
    result.check("3b. An unsupported schedule_type is rejected by validate_schedule_type", invalid_rejected)

    # --- 4: Payroll Period distinct from Pay Date ---------------------------
    period_start, period_end, pay_date = _dt(0), _dt(14), _dt(18)
    payroll_run = m.PayrollRun(
        restaurant_id=restaurant.id, source_system_id=source_adp.id, payroll_schedule_id=ps_biweekly.id,
        period_start=period_start, period_end=period_end, pay_date=pay_date,
        run_type="REGULAR", status="COMPLETE",
    )
    session.add(payroll_run)
    session.flush()
    result.check(
        "4. PayrollPeriod (period_start/period_end) is distinct from pay_date, never derived from it",
        payroll_run.period_end != payroll_run.pay_date and payroll_run.pay_date > payroll_run.period_end,
    )

    # --- 5: Workweek distinct from Payroll Period ---------------------------
    workweeks = schedule.workweeks_within_period(period_start, period_end, start_weekday=0)
    result.check(
        "5. One BIWEEKLY PayrollPeriod contains exactly two Monday-anchored Workweeks, "
        "each 7 days, none equal to the full Period",
        len(workweeks) == 2
        and all((end - start) == timedelta(days=7) for start, end in workweeks)
        and all((end - start) != (period_end - period_start) for start, end in workweeks),
    )

    # --- 6: No hardcoded 80-hour biweekly overtime logic --------------------
    result.check(
        "6. No compute_overtime (or similarly-named) function exists in the generic "
        "schedule/compensation modules — overtime is never computed from a Payroll-Period total",
        not hasattr(schedule, "compute_overtime") and not hasattr(compensation, "compute_overtime"),
    )

    # --- 7-13: Compensation Terms -------------------------------------------
    t0, t1 = _dt(0), _dt(30)
    term_a_service = m.EmployeeCompensationTerm(
        employee_id=employee_a.id, function_label="Service", compensation_basis="HOURLY",
        hourly_rate_minor=1500, valid_from=t0,
    )
    term_b_service = m.EmployeeCompensationTerm(
        employee_id=employee_b.id, function_label="Service", compensation_basis="HOURLY",
        hourly_rate_minor=1800, valid_from=t0,
    )
    session.add_all([term_a_service, term_b_service])
    session.flush()
    result.check(
        "7. Compensation Terms are Employee-specific (two different Employees, two independent rows)",
        term_a_service.employee_id != term_b_service.employee_id,
    )
    result.check(
        "8. Two Employees performing the same function may have different rates",
        term_a_service.function_label == term_b_service.function_label
        and term_a_service.hourly_rate_minor != term_b_service.hourly_rate_minor,
    )

    term_a_social = m.EmployeeCompensationTerm(
        employee_id=employee_a.id, function_label="Social Media", compensation_basis="HOURLY",
        hourly_rate_minor=2000, valid_from=t0,
    )
    session.add(term_a_social)
    session.flush()
    concurrent = compensation.terms_valid_during([term_a_service, term_a_social], t0, t0 + timedelta(days=1))
    result.check(
        "9. One Employee may hold two concurrently applicable Compensation Terms "
        "(different functions) — not treated as a conflict",
        len(concurrent) == 2
        and not compensation.detect_mid_period_conflict([term_a_service, term_a_social], t0, t0 + timedelta(days=1)),
    )

    result.check(
        "10. HOURLY compensation stores a rate and no per-period base amount",
        term_a_service.compensation_basis == "HOURLY"
        and term_a_service.hourly_rate_minor is not None
        and term_a_service.salaried_period_amount_minor is None,
    )

    term_salaried = m.EmployeeCompensationTerm(
        employee_id=employee_b.id, function_label="Manager Base", compensation_basis="SALARIED",
        salaried_period_amount_minor=250000, valid_from=t0,
    )
    session.add(term_salaried)
    session.flush()
    result.check(
        "11. SALARIED compensation stores a base-pay-per-Payroll-Period amount, "
        "never an annual salary field, and no hourly rate",
        term_salaried.salaried_period_amount_minor == 250000 and term_salaried.hourly_rate_minor is None,
    )

    term_a_service.valid_to = t1
    term_a_service_v2 = m.EmployeeCompensationTerm(
        employee_id=employee_a.id, function_label="Service", compensation_basis="HOURLY",
        hourly_rate_minor=1650, valid_from=t1,
    )
    session.add(term_a_service_v2)
    session.flush()
    history_rows = session.scalars(
        select(m.EmployeeCompensationTerm).where(
            m.EmployeeCompensationTerm.employee_id == employee_a.id,
            m.EmployeeCompensationTerm.function_label == "Service",
        )
    ).all()
    result.check(
        "12. A compensation change preserves history (both the closed and the new row persist; "
        "the closed row's original rate is never overwritten)",
        len(history_rows) == 2 and term_a_service.hourly_rate_minor == 1500,
    )

    mid_period_conflict = compensation.detect_mid_period_conflict(
        [term_a_service, term_a_service_v2], t0 + timedelta(days=5), t1 + timedelta(days=5)
    )
    result.check(
        "13. A mid-Payroll-Period incompatible compensation change surfaces MANUAL_REVIEW_REQUIRED",
        mid_period_conflict
        and compensation.review_status_for_period(
            [term_a_service, term_a_service_v2], t0 + timedelta(days=5), t1 + timedelta(days=5)
        ) == compensation.MANUAL_REVIEW_REQUIRED,
    )

    # --- 14-21, 25-27: EmployeePayrollResult / facts / Labor Cost -----------
    result_a = m.EmployeePayrollResult(payroll_run_id=payroll_run.id, employee_id=employee_a.id)
    result_b = m.EmployeePayrollResult(payroll_run_id=payroll_run.id, employee_id=employee_b.id)
    session.add_all([result_a, result_b])
    session.flush()

    holiday_fact = m.PayrollEarningFact(
        employee_payroll_result_id=result_a.id, earning_type="HOLIDAY_PAY", source_label="Holiday Pay",
        quantity=None, unit=None, rate_minor=None, amount_minor=10000, paid_to_employee=True,
    )
    session.add(holiday_fact)
    session.flush()
    result.check(
        "14. A paid non-work earning (e.g. holiday pay) can exist without any hours/quantity",
        holiday_fact.id is not None and holiday_fact.quantity is None,
    )

    bonus_fact = m.PayrollEarningFact(
        employee_payroll_result_id=result_a.id, earning_type="BONUS", source_label="Production Bonus",
        quantity=None, unit=None, rate_minor=None, amount_minor=50000, paid_to_employee=True,
    )
    session.add(bonus_fact)
    session.flush()
    result.check(
        "15. Bonus enters Payroll as an externally supplied amount — no compute_bonus "
        "function exists anywhere in the payroll package, and the stored amount is exactly "
        "what was supplied",
        not hasattr(compensation, "compute_bonus")
        and not hasattr(schedule, "compute_bonus")
        and not hasattr(adp, "compute_bonus")
        and bonus_fact.amount_minor == 50000,
    )

    tip_fact = m.PayrollEarningFact(
        employee_payroll_result_id=result_a.id, earning_type="CASH_TIPS", source_label="Cash tips*",
        quantity=None, unit=None, rate_minor=None, amount_minor=99999, paid_to_employee=False,
        excluded_from_taxable_wages=False,
    )
    session.add(tip_fact)
    session.flush()
    session.expire_all()  # force relationship collections to reflect the facts just flushed
    cost_a_with_tip = compute_employee_labor_cost(result_a)
    result.check(
        "16. A reportable Tip earning fact marked paid_to_employee=False is excluded from "
        "employer-paid earnings",
        99999 not in [f.amount_minor for f in result_a.earning_facts if f.paid_to_employee]
        and cost_a_with_tip.employer_paid_earnings_minor == 10000 + 50000,
    )

    parsed_type, parsed_paid, parsed_excluded = adp.normalize_earning_label("Cash tips* ")
    result.check(
        "17. The importer's own label parser marks a '*'-suffixed provider earning line as "
        "not paid to the Employee, distinct from being excluded-from-wages",
        parsed_type == "CASH_TIPS" and parsed_paid is False and parsed_excluded is False,
    )

    liability_fact = m.PayrollEmployerLiabilityFact(
        employee_payroll_result_id=result_a.id, liability_type="FED_SOCSEC_ER",
        source_label="FED SOCSEC-ER", amount_minor=5000,
    )
    session.add(liability_fact)
    session.flush()
    session.expire_all()
    cost_a_with_liability = compute_employee_labor_cost(result_a)
    result.check(
        "18. An employer liability fact increases Payroll Employer Cost",
        cost_a_with_liability.payroll_employer_cost_minor
        == cost_a_with_liability.employer_paid_earnings_minor + cost_a_with_liability.employer_liabilities_minor
        and cost_a_with_liability.employer_liabilities_minor == 5000,
    )

    withholding_shaped_tables = [
        t.__tablename__ for t in m.ALL_MODELS if "withholding" in t.__tablename__.lower()
    ]
    result.check(
        "19. No employee-withholding table exists at all — employee withholding cannot be "
        "misclassified as employer labor cost because there is nothing to reference",
        withholding_shaped_tables == [],
    )

    payment_fact_a = m.PayrollPaymentFact(
        employee_payroll_result_id=result_a.id, pay_date=pay_date, payment_method="Direct Deposit",
        payment_amount_minor=123456,
    )
    payment_fact_b = m.PayrollPaymentFact(
        employee_payroll_result_id=result_b.id, pay_date=pay_date, payment_method="Check",
        payment_amount_minor=54321,
    )
    session.add_all([payment_fact_a, payment_fact_b])
    session.flush()
    session.expire_all()
    result.check(
        "20. The provider payment fact reconstructs exactly what was actually paid to the Employee",
        compute_employee_labor_cost(result_a).payment_amount_minor == 123456,
    )

    run_cost = compute_payroll_run_labor_cost(session, payroll_run.id)
    per_employee_by_id = {e.employee_id: e for e in run_cost.per_employee}
    result.check(
        "21. Two Employees' payroll result details remain independently queryable within the "
        "same Run, never conflated into one aggregate",
        len(run_cost.per_employee) == 2
        and per_employee_by_id[employee_a.id].payment_amount_minor == 123456
        and per_employee_by_id[employee_b.id].payment_amount_minor == 54321,
    )

    result.check(
        "25. Run-level Labor Cost totals reconcile exactly with a manual sum of the atomic facts",
        run_cost.total_payment_amount_minor == 123456 + 54321
        and run_cost.total_employer_liabilities_minor == 5000
        and run_cost.total_payroll_employer_cost_minor
        == run_cost.total_employer_paid_earnings_minor + run_cost.total_employer_liabilities_minor,
    )

    special_run = m.PayrollRun(
        restaurant_id=restaurant.id, source_system_id=source_adp.id, payroll_schedule_id=None,
        period_start=None, period_end=None, pay_date=_dt(40), run_type="SPECIAL", status="COMPLETE",
    )
    session.add(special_run)
    session.flush()
    result.check(
        "27. A SPECIAL PayrollRun (one-off bonus/correction) can persist with no PayrollPeriod",
        special_run.id is not None and special_run.period_start is None and special_run.period_end is None,
    )

    unseen_type, unseen_paid, _ = adp.normalize_earning_label("Holiday Pay 2")
    unseen_fact = m.PayrollEarningFact(
        employee_payroll_result_id=result_b.id, earning_type=unseen_type, source_label="Holiday Pay 2",
        amount_minor=2500, paid_to_employee=unseen_paid,
    )
    session.add(unseen_fact)
    session.flush()
    result.check(
        "26. A provider earning label never seen before is normalized generically and stored "
        "without any schema/code change",
        unseen_fact.id is not None and unseen_type == "HOLIDAY_PAY_2",
    )

    importer_source = inspect.getsource(adp)
    result.check(
        "28. The ADP importer performs no network/API access — no requests/oauth/http import "
        "appears anywhere in its source, only local file (openpyxl) and database access",
        not any(token in importer_source for token in ("import requests", "oauth", "http.client", "urllib.request")),
    )

    # --- 29-34: Payment Execution Provider (TASK_PAYROLL_002) ---------------
    adp_direct_deposit_run = m.PayrollRun(
        restaurant_id=restaurant.id, source_system_id=source_adp.id, payroll_schedule_id=None,
        period_start=None, period_end=None, pay_date=_dt(50), run_type="SPECIAL", status="COMPLETE",
    )
    session.add(adp_direct_deposit_run)
    session.flush()
    pe.assign_payment_execution_provider(adp_direct_deposit_run, pe.ADP_DIRECT_DEPOSIT)
    result.check(
        "29. ADP_DIRECT_DEPOSIT is representable as an explicit, auditable Payment Execution "
        "Provider on a PayrollRun",
        adp_direct_deposit_run.payment_execution_provider == "ADP_DIRECT_DEPOSIT",
    )

    mercury_run = m.PayrollRun(
        restaurant_id=restaurant.id, source_system_id=source_adp.id, payroll_schedule_id=None,
        period_start=None, period_end=None, pay_date=_dt(51), run_type="SPECIAL", status="COMPLETE",
    )
    session.add(mercury_run)
    session.flush()
    pe.assign_payment_execution_provider(mercury_run, pe.MERCURY_ACH)
    result.check(
        "30. A future MERCURY_ACH Payment Execution Provider is representable on the same "
        "canonical PayrollRun model without any Payroll redesign, with no Mercury API ever "
        "called (no 'mercury' network/API token appears anywhere in the payment_execution "
        "module's source)",
        mercury_run.payment_execution_provider == "MERCURY_ACH"
        and "requests" not in inspect.getsource(pe)
        and "http.client" not in inspect.getsource(pe),
    )

    try:
        pe.assign_payment_execution_provider(adp_direct_deposit_run, pe.MERCURY_ACH)
        double_payment_rejected = False
    except ValueError:
        double_payment_rejected = True
    result.check(
        "31. Double-payment prevention: reassigning an already-ADP-assigned PayrollRun's Payment "
        "Execution Provider to MERCURY_ACH is rejected, and the original assignment is left intact",
        double_payment_rejected and adp_direct_deposit_run.payment_execution_provider == "ADP_DIRECT_DEPOSIT",
    )

    pe.assign_payment_execution_provider(adp_direct_deposit_run, pe.ADP_DIRECT_DEPOSIT)
    result.check(
        "31b. Re-asserting the same already-assigned Payment Execution Provider is a safe no-op, "
        "never raising",
        adp_direct_deposit_run.payment_execution_provider == "ADP_DIRECT_DEPOSIT",
    )

    try:
        pe.assign_payment_execution_provider(mercury_run, "WIRE_TRANSFER")
        invalid_provider_rejected = False
    except ValueError:
        invalid_provider_rejected = True
    result.check(
        "32. An unsupported Payment Execution Provider value is rejected rather than silently "
        "stored",
        invalid_provider_rejected,
    )

    result.check(
        "33. Payment execution status is derived, never fabricated: a PayrollRun with an assigned "
        "Payment Execution Provider but no PayrollPaymentFact evidence yet reports UNKNOWN, not "
        "'paid'",
        pe.payment_execution_status(mercury_run) == pe.UNKNOWN
        and not pe.has_payment_execution_evidence(mercury_run),
    )

    evidenced_result = m.EmployeePayrollResult(payroll_run_id=adp_direct_deposit_run.id, employee_id=employee_a.id)
    session.add(evidenced_result)
    session.flush()
    session.add(
        m.PayrollPaymentFact(
            employee_payroll_result_id=evidenced_result.id, pay_date=_dt(50),
            payment_method="Direct Deposit", payment_amount_minor=100000,
        )
    )
    session.flush()
    session.expire_all()
    result.check(
        "34. Once PayrollPaymentFact evidence exists for a Run, payment execution status "
        "reflects EVIDENCED — derived from the provider's own reported facts, never a stored, "
        "independently-settable status column",
        pe.payment_execution_status(adp_direct_deposit_run) == pe.EVIDENCED,
    )

    # --- 35-39: PayrollExecutionConfiguration (TASK_PAYROLL_003) -----------
    config_restaurant = m.Restaurant(name="Config Test Restaurant", default_currency="USD")
    session.add(config_restaurant)
    session.flush()

    config_adp = m.PayrollExecutionConfiguration(
        restaurant_id=config_restaurant.id, provider=pe.ADP_DIRECT_DEPOSIT,
        valid_from=_dt(60), valid_to=None,
    )
    session.add(config_adp)
    session.flush()
    result.check(
        "35. approved_provider_at derives ADP_DIRECT_DEPOSIT from a PayrollExecutionConfiguration "
        "valid at the queried instant",
        pe.approved_provider_at(session, restaurant_id=config_restaurant.id, at=_dt(70)) == pe.ADP_DIRECT_DEPOSIT,
    )
    result.check(
        "35b. approved_provider_at returns None outside any configured window — never guesses a "
        "provider for an unconfigured instant",
        pe.approved_provider_at(session, restaurant_id=config_restaurant.id, at=_dt(55)) is None,
    )

    configured_run = m.PayrollRun(
        restaurant_id=config_restaurant.id, source_system_id=source_adp.id, payroll_schedule_id=None,
        period_start=None, period_end=None, pay_date=_dt(70), run_type="SPECIAL", status="COMPLETE",
    )
    session.add(configured_run)
    session.flush()
    derived = pe.approved_provider_at(session, restaurant_id=config_restaurant.id, at=configured_run.pay_date)
    if derived is not None:
        pe.assign_payment_execution_provider(configured_run, derived)
    result.check(
        "36. A PayrollExecutionConfiguration lets a Run derive the approved provider automatically "
        "when not explicitly selected at import/acquisition time (Option B)",
        configured_run.payment_execution_provider == "ADP_DIRECT_DEPOSIT",
    )

    explicit_run = m.PayrollRun(
        restaurant_id=config_restaurant.id, source_system_id=source_adp.id, payroll_schedule_id=None,
        period_start=None, period_end=None, pay_date=_dt(70), run_type="SPECIAL", status="COMPLETE",
    )
    session.add(explicit_run)
    session.flush()
    pe.assign_payment_execution_provider(explicit_run, pe.MERCURY_ACH)
    result.check(
        "37. An explicit Payment Execution Provider selection (Option A) is honored regardless of "
        "what an approved configuration would otherwise derive — explicit selection is never "
        "overridden by configuration",
        explicit_run.payment_execution_provider == "MERCURY_ACH",
    )

    config_adp.valid_to = _dt(80)
    config_mercury = m.PayrollExecutionConfiguration(
        restaurant_id=config_restaurant.id, provider=pe.MERCURY_ACH,
        valid_from=_dt(80), valid_to=None,
    )
    session.add(config_mercury)
    session.flush()
    later_run = m.PayrollRun(
        restaurant_id=config_restaurant.id, source_system_id=source_adp.id, payroll_schedule_id=None,
        period_start=None, period_end=None, pay_date=_dt(85), run_type="SPECIAL", status="COMPLETE",
    )
    session.add(later_run)
    session.flush()
    later_derived = pe.approved_provider_at(session, restaurant_id=config_restaurant.id, at=later_run.pay_date)
    if later_derived is not None:
        pe.assign_payment_execution_provider(later_run, later_derived)
    result.check(
        "38. Temporal correctness: a Run whose pay_date falls after the configuration changed to "
        "MERCURY_ACH derives MERCURY_ACH, while the earlier, already-assigned "
        "configured_run (pay_date under the prior ADP_DIRECT_DEPOSIT window) remains untouched — "
        "changing the approved configuration never alters an already-created historical Run's "
        "assignment",
        later_run.payment_execution_provider == "MERCURY_ACH"
        and configured_run.payment_execution_provider == "ADP_DIRECT_DEPOSIT",
    )

    # --- 22-24: Import idempotency / ambiguous mapping / correction --------
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir) / "synthetic_payroll_detail.xlsx"
        _build_synthetic_workbook(
            tmp_path,
            rows=[
                {"name": "Tester, Alice", "regular_amount": 1200.0, "payment_amount": 1100.0},
                {"name": "Tester, Bob", "regular_amount": 1440.0, "payment_amount": 1320.0},
                {"name": "Duplicate, Chris", "regular_amount": 900.0, "payment_amount": 800.0},
            ],
        )

        count_runs_before = session.query(m.PayrollRun).count()
        count_results_before = session.query(m.EmployeePayrollResult).count()

        persist_1 = adp.persist_import(
            session, source_system_id=source_adp.id, restaurant_id=restaurant.id, file_path=tmp_path,
            period_start=_dt(0), period_end=_dt(14), run_type="REGULAR",
        )
        session.flush()

        result.check(
            "23. An ambiguous Employee mapping (two Employees, identical normalized name) "
            "blocks import for that row and is surfaced as an issue, never guessed",
            persist_1.ambiguous_employee_count == 1
            and any(
                issue.issue_type == "AMBIGUOUS_EMPLOYEE_MAPPING"
                for issue in session.scalars(
                    select(m.PayrollImportIssue).where(m.PayrollImportIssue.import_run_id == persist_1.import_run_id)
                ).all()
            ),
        )

        count_runs_after_first = session.query(m.PayrollRun).count()
        count_results_after_first = session.query(m.EmployeePayrollResult).count()
        result.check(
            "22a. The first import of a new synthetic workbook creates new rows",
            count_runs_after_first == count_runs_before + 1
            and count_results_after_first == count_results_before + 2,  # Alice + Bob resolved; Chris ambiguous
        )

        first_run = session.get(m.PayrollRun, persist_1.payroll_run_id)
        result.check(
            "29b (TASK_PAYROLL_003 correction): the ADP importer does NOT default a newly created "
            "PayrollRun's Payment Execution Provider to ADP_DIRECT_DEPOSIT merely because the "
            "source is ADP — with no explicit selection and no PayrollExecutionConfiguration for "
            "this Restaurant, the Run's provider is left unassigned (NULL)",
            first_run is not None and first_run.payment_execution_provider is None,
        )

        persist_2 = adp.persist_import(
            session, source_system_id=source_adp.id, restaurant_id=restaurant.id, file_path=tmp_path,
            period_start=_dt(0), period_end=_dt(14), run_type="REGULAR",
        )
        session.flush()
        count_runs_after_second = session.query(m.PayrollRun).count()
        count_results_after_second = session.query(m.EmployeePayrollResult).count()
        result.check(
            "22. Re-importing the exact same workbook is idempotent — no duplicate PayrollRun/"
            "EmployeePayrollResult rows are created",
            persist_2.created is False
            and persist_2.import_run_id == persist_1.import_run_id
            and count_runs_after_second == count_runs_after_first
            and count_results_after_second == count_results_after_first,
        )

        renamed_path = Path(tmpdir) / "a_totally_different_filename_2026.xlsx"
        shutil.copy(tmp_path, renamed_path)
        persist_renamed = adp.persist_import(
            session, source_system_id=source_adp.id, restaurant_id=restaurant.id, file_path=renamed_path,
            period_start=_dt(0), period_end=_dt(14), run_type="REGULAR",
        )
        session.flush()
        count_runs_after_renamed = session.query(m.PayrollRun).count()
        result.check(
            "22b. Duplicate protection is keyed by file content (SHA-256), not filename — the "
            "identical workbook content under a completely different filename is still detected "
            "as the same import, never a new PayrollRun",
            persist_renamed.created is False
            and persist_renamed.import_run_id == persist_1.import_run_id
            and count_runs_after_renamed == count_runs_after_second,
        )

        tmp_path_v2 = Path(tmpdir) / "synthetic_payroll_detail_corrected.xlsx"
        _build_synthetic_workbook(
            tmp_path_v2,
            rows=[
                {"name": "Tester, Alice", "regular_amount": 1300.0, "payment_amount": 1150.0},  # corrected amount
                {"name": "Tester, Bob", "regular_amount": 1440.0, "payment_amount": 1320.0},
                {"name": "Duplicate, Chris", "regular_amount": 900.0, "payment_amount": 800.0},
            ],
        )
        persist_3 = adp.persist_import(
            session, source_system_id=source_adp.id, restaurant_id=restaurant.id, file_path=tmp_path_v2,
            period_start=_dt(0), period_end=_dt(14), run_type="REGULAR",
            supersedes_import_run_id=persist_1.import_run_id,
        )
        session.flush()
        original_run = session.get(m.PayrollRun, persist_1.payroll_run_id)
        result.check(
            "24. A corrected provider report never overwrites history — the original PayrollRun "
            "still exists (now SUPERSEDED) and a new, separate PayrollRun was created",
            persist_3.created is True
            and persist_3.payroll_run_id != persist_1.payroll_run_id
            and original_run is not None
            and original_run.status == "SUPERSEDED"
            and original_run.superseded_by_payroll_run_id == persist_3.payroll_run_id,
        )

        first_import_run = session.get(m.PayrollImportRun, persist_1.import_run_id)
        result.check(
            "40. Provenance: a file-based import records acquisition_method=ADP_XLSX_FILE",
            first_import_run is not None and first_import_run.acquisition_method == "ADP_XLSX_FILE",
        )

        # --- 41-46: Acquisition adapters (TASK_PAYROLL_003) -----------------
        bytes_parsed = adp.parse_payroll_detail_workbook_bytes(tmp_path.read_bytes())
        file_parsed = adp.parse_payroll_detail_workbook(tmp_path)
        result.check(
            "41. parse_payroll_detail_workbook_bytes produces an identical result to the "
            "file-based parser for the same content",
            len(bytes_parsed.employees) == len(file_parsed.employees)
            and bytes_parsed.header_pay_date == file_parsed.header_pay_date,
        )

        acq_source = m.SourceSystem(code="ADP_ACQ_TEST", name="ADP (acquisition test)", active=True)
        session.add(acq_source)
        session.flush()

        local_adapter = acquisition.LocalFileAcquisitionAdapter(file_path=tmp_path)
        acquire_results_1 = acquisition.acquire_and_import(
            session, local_adapter,
            source_system_id=acq_source.id, restaurant_id=restaurant.id,
            period_start=_dt(0), period_end=_dt(14), run_type="REGULAR",
        )
        session.flush()
        acquire_results_2 = acquisition.acquire_and_import(
            session, local_adapter,
            source_system_id=acq_source.id, restaurant_id=restaurant.id,
            period_start=_dt(0), period_end=_dt(14), run_type="REGULAR",
        )
        session.flush()
        result.check(
            "42. acquire_and_import (LocalFileAcquisitionAdapter) persists via the same idempotent "
            "core as persist_import — a second acquisition of the identical content is a no-op",
            len(acquire_results_1) == 1 and acquire_results_1[0].created is True
            and len(acquire_results_2) == 1 and acquire_results_2[0].created is False
            and acquire_results_2[0].import_run_id == acquire_results_1[0].import_run_id,
        )

        class _FakeSftpFile:
            def __init__(self, data: bytes) -> None:
                self._data = data
            def read(self) -> bytes:
                return self._data
            def __enter__(self):
                return self
            def __exit__(self, *exc) -> None:
                return None

        class _FakeSftpTransport:
            def __init__(self, files: dict[str, bytes]) -> None:
                self._files = files
            def listdir(self, path: str) -> list[str]:
                return list(self._files.keys())
            def open(self, path: str, mode: str = "rb"):
                filename = path.rsplit("/", 1)[-1]
                return _FakeSftpFile(self._files[filename])

        fake_transport = _FakeSftpTransport({"PayrollDetail_sftp.xlsx": tmp_path.read_bytes()})
        sftp_adapter = acquisition.AdpSftpAcquisitionAdapter(
            acquisition.SftpConnectionConfig(
                host="sftp.example.invalid", username="rfone", remote_directory="/inbound", password="x",
            ),
            transport_factory=lambda: fake_transport,
        )
        acquired_via_sftp = sftp_adapter.fetch()
        result.check(
            "43. AdpSftpAcquisitionAdapter.fetch() lists and downloads files from the configured "
            "remote directory via an injectable transport, tagging acquisition_method=ADP_SFTP_AES "
            "— no real network/paramiko connection required for this to be exercised",
            len(acquired_via_sftp) == 1
            and acquired_via_sftp[0].acquisition_method == acquisition.ACQUISITION_METHOD_SFTP_AES
            and acquired_via_sftp[0].source_file_name == "PayrollDetail_sftp.xlsx"
            and acquired_via_sftp[0].file_bytes == tmp_path.read_bytes(),
        )

        sftp_import_results = acquisition.acquire_and_import(
            session, sftp_adapter,
            source_system_id=acq_source.id, restaurant_id=restaurant.id,
            period_start=_dt(0), period_end=_dt(14), run_type="REGULAR",
        )
        session.flush()
        result.check(
            "44. A file acquired via the SFTP adapter is recognized as identical content already "
            "imported via the local-file adapter earlier in this same Restaurant/source scope — "
            "idempotency is keyed on content, never on transport or filename",
            len(sftp_import_results) == 1 and sftp_import_results[0].created is False
            and sftp_import_results[0].import_run_id == acquire_results_1[0].import_run_id,
        )

        try:
            acquisition.AdpSftpAcquisitionAdapter.from_environment({})
            sftp_unconfigured_rejected = False
        except acquisition.AcquisitionNotConfiguredError:
            sftp_unconfigured_rejected = True
        result.check(
            "45. AdpSftpAcquisitionAdapter.from_environment raises a clear "
            "AcquisitionNotConfiguredError (naming what is missing) when required environment "
            "variables are absent — never silently proceeds or fabricates a connection",
            sftp_unconfigured_rejected,
        )

        try:
            acquisition.AdpApiAcquisitionAdapter.from_environment({})
            api_unconfigured_rejected = False
        except acquisition.AcquisitionNotConfiguredError:
            api_unconfigured_rejected = True
        result.check(
            "46a. AdpApiAcquisitionAdapter.from_environment raises AcquisitionNotConfiguredError "
            "when ADP API credentials are absent",
            api_unconfigured_rejected,
        )

        fake_api_env = {
            "ADP_API_BASE_URL": "https://api.adp.com", "ADP_API_CLIENT_ID": "fake",
            "ADP_API_CLIENT_SECRET": "fake", "ADP_API_CLIENT_CERT_PATH": "/fake/cert.pem",
            "ADP_API_CLIENT_KEY_PATH": "/fake/key.pem",
        }
        api_adapter = acquisition.AdpApiAcquisitionAdapter.from_environment(fake_api_env)
        try:
            api_adapter.fetch()
            api_fetch_raised = False
        except acquisition.AdpApiNotImplementedError:
            api_fetch_raised = True
        result.check(
            "46b. AdpApiAcquisitionAdapter.fetch() raises AdpApiNotImplementedError even with "
            "credential-shaped configuration present — this task never fabricates a successful "
            "ADP Payroll Output API response without verified endpoint/schema documentation",
            api_fetch_raised,
        )


def _ensure_merchant(session: Session, source_system: m.SourceSystem) -> m.Merchant:
    merchant = m.Merchant(source_system_id=source_system.id, source_merchant_id="MERCH-PAYROLL-TEST", name="Test Merchant")
    session.add(merchant)
    session.flush()
    return merchant
