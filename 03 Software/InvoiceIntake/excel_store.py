"""
Persistence layer: appends reviewed Purchase Documents and Purchase Lines to
a single Excel workbook, so the data can be opened and checked directly
(this is a prototype -- Excel stands in for a real database for now).

Column names follow 00 Knowledge Repository/Domains/Restaurant/Purchasing/DataDictionary.md
for the Purchase Document and Purchase Line entities. A few practical columns
(SourceFile, RawText) are added for traceability, consistent with the Core
Traceability principle. Cost-normalization columns (NormalizedQuantity,
CostPerGram, etc.) are left as empty headers, ready for a later step.
"""
import os
import threading
from datetime import datetime

import openpyxl
from openpyxl import Workbook

DOC_HEADERS = [
    "PurchaseDocumentId",
    "SupplierName",
    "DocumentNumber",
    "DocumentType",
    "IssueDate",
    "AcquisitionMethod",
    "Currency",
    "TotalAmount",
    "Status",
    "SourceFile",
    "CreatedAt",
]

LINE_HEADERS = [
    "PurchaseLineId",
    "PurchaseDocumentId",
    "SupplierDescription",
    "Quantity",
    "PurchaseUnit",
    "UnitPrice",
    "LineAmount",
    "NormalizedQuantity",
    "CostPerGram",
]

_lock = threading.Lock()


def _ensure_workbook(path: str) -> Workbook:
    if os.path.exists(path):
        return openpyxl.load_workbook(path)
    wb = Workbook()
    docs_ws = wb.active
    docs_ws.title = "PurchaseDocuments"
    docs_ws.append(DOC_HEADERS)
    lines_ws = wb.create_sheet("PurchaseLines")
    lines_ws.append(LINE_HEADERS)
    return wb


def _next_id(ws, id_col_index: int = 1) -> int:
    max_id = 0
    for row in ws.iter_rows(min_row=2, max_col=id_col_index, values_only=True):
        value = row[id_col_index - 1]
        if isinstance(value, int):
            max_id = max(max_id, value)
    return max_id + 1


def save_purchase_document(path: str, header: dict, lines: list[dict], source_file: str) -> int:
    """Appends one Purchase Document + its Purchase Lines. Returns the new PurchaseDocumentId."""
    with _lock:
        wb = _ensure_workbook(path)
        docs_ws = wb["PurchaseDocuments"]
        lines_ws = wb["PurchaseLines"]

        doc_id = _next_id(docs_ws)
        docs_ws.append(
            [
                doc_id,
                header.get("supplier_name", ""),
                header.get("document_number", ""),
                header.get("document_type", "Invoice"),
                header.get("issue_date", ""),
                header.get("acquisition_method", "OCR"),
                header.get("currency", ""),
                _to_number(header.get("total_amount", "")),
                "Reviewed",
                source_file,
                datetime.now().isoformat(timespec="seconds"),
            ]
        )

        line_id = _next_id(lines_ws)
        for line in lines:
            if not any(line.get(k) for k in ("description", "quantity", "unit_price", "line_amount")):
                continue
            lines_ws.append(
                [
                    line_id,
                    doc_id,
                    line.get("description", ""),
                    _to_number(line.get("quantity", "")),
                    line.get("unit", ""),
                    _to_number(line.get("unit_price", "")),
                    _to_number(line.get("line_amount", "")),
                    "",
                    "",
                ]
            )
            line_id += 1

        wb.save(path)
        return doc_id


def _to_number(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return value
